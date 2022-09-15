import xlwt
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sqlExcute

#创建完整信息表
def writeToExcel(dateList):
    # 创建新的workbook（其实就是创建新的excel）
    workbook = xlwt.Workbook(encoding='ascii')

    j = 0
    while j < len(dateList):
        worksheet = workbook.add_sheet(dateList[j])# 创建新的sheet表

        # 合并 第0行到第2行 的 第0列到第3列
        worksheet.write_merge(0, 2, 0, 4, dateList[j] + "疫情通报表")
        worksheet.write_merge(0, 0, 5, 7, '当日新增本地确诊')
        worksheet.write_merge(2, 2, 5, 7, '当日新增本地无症状感染者')

        # 往表格写入内容
        worksheet.write(3, 0, "省份")
        worksheet.write(3, 1, "本土新增")
        worksheet.write(3, 2, "本土新增无症状感染者")
        # 获取每日新增并写入
        newAdd = sqlExcute.selectNewAdd(dateList[j])
        worksheet.write(0,8,newAdd)
        # 获取每日新增无症状感染者并写入
        newAsymptomatic = sqlExcute.selectNewAsymptomatic(dateList[j])
        worksheet.write(2,8,newAsymptomatic)
        #写入省份信息
        i = 4
        provinceList = sqlExcute.selectProvince()
        print(dateList[j])
        for province in provinceList:
            worksheet.write(i, 0, province)#省份写入
            #对港澳台进行判断通过当天减前一天得出每日新增
            if(j < len(dateList) - 1 and (re.findall('\'(.*)\',', str(province))[0] == '香港' or re.findall('\'(.*)\',', str(province))[0] == '台湾' or re.findall('\'(.*)\',', str(province))[0] == '澳门')):
                #当日累计
                provinceNewAddToday = sqlExcute.selectProvinceNewAdd(dateList[j],re.findall('\'(.*)\',',str(province))[0])
                #港澳台无症状感染者无相关信息故都为0
                provinceNewAsymptomatic = sqlExcute.selectProvinceAsymptomatic(dateList[j],re.findall('\'(.*)\',',str(province))[0])
                #昨日累计
                provinceNewAddYes = sqlExcute.selectProvinceNewAdd(dateList[j+1],re.findall('\'(.*)\',', str(province))[0])
                #格式转化
                provinceNewAddToday = re.findall('\'([0-9]+)\'',str(provinceNewAddToday))[0]
                provinceNewAddYes = re.findall('\'([0-9]+)\'',str(provinceNewAddYes))[0]
                #当日新增
                provinceNewAdd = int(provinceNewAddToday) - int(provinceNewAddYes)
                # 写入
                worksheet.write(i,1,str(provinceNewAdd))
                worksheet.write(i,2,provinceNewAsymptomatic)
            else:
                provinceNewAdd = sqlExcute.selectProvinceNewAdd(dateList[j], re.findall('\'(.*)\',', str(province))[0])
                provinceNewAsymptomatic = sqlExcute.selectProvinceAsymptomatic(dateList[j],re.findall('\'(.*)\',', str(province))[0])
                worksheet.write(i, 1, provinceNewAdd)
                worksheet.write(i, 2, provinceNewAsymptomatic)
            i += 1
        j+=1
    # 保存
    workbook.save("疫情通报表1.1.xls")

#创建全国每日信息表
def writeCountryInfo(dateList):
    # 创建新的workbook（其实就是创建新的excel）
    workbook = xlwt.Workbook(encoding='ascii')
    # 创建新的sheet表
    worksheet = workbook.add_sheet('sheet1')
    # 插入表中基本信息
    worksheet.write(0, 0, "日期")
    worksheet.write(0, 1, "新增确诊")
    worksheet.write(0, 2, "新增无症状感染者")
    # 遍历插入每日信息
    i = 1
    for date in dateList:
        newAsymptomatic = sqlExcute.selectNewAsymptomatic(date)
        newAdd = sqlExcute.selectNewAdd(date)
        worksheet.write(i, 0, date)
        worksheet.write(i, 1, newAdd)
        worksheet.write(i, 2, newAsymptomatic)
        i += 1
    # 保存
    workbook.save("全国疫情通报表1.1.xls")

dateList = sqlExcute.selectFromDateInfo();
# writeToExcel(dateList)
# writeCountryInfo(dateList)